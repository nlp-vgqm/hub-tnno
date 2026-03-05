'''
- 完整的对话系统架构

- NLU（自然语言理解）：负责意图识别和槽位提取
- DST（对话状态追踪）：管理对话状态和槽位填充
- PM（策略管理）：处理对话流程和决策
- NLG（自然语言生成）：生成系统响应
- 本地大模型集成

- 使用ollama库调用本地qwen3-14B模型
- 当大模型调用失败时，自动回退到传统方法
- 增强功能

- 重听功能 ：用户可以请求系统重复上一条回答
- 推荐功能 ：当用户请求推荐时，系统返回槽位的可选值
- 确认跳转 ：用户确认后自动跳转到下一个对话节点
- 场景支持

- 买衣服场景：支持选择服装类型、颜色、尺寸，以及分期付款
- 看电影场景：支持选择时间和电影名称
- 槽位提取优化

- 基于Excel配置的槽位值和反问方式
- 支持模糊匹配和相似度计算
- 处理用户自由文本输入
'''



import json
import re
import ollama
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple


def levenshtein_distance(a: str, b: str) -> int:
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev_row = list(range(len(b) + 1))
    for i, ca in enumerate(a, start=1):
        curr_row = [i]
        for j, cb in enumerate(b, start=1):
            ins = curr_row[j - 1] + 1
            delete = prev_row[j] + 1
            replace = prev_row[j - 1] + (0 if ca == cb else 1)
            curr_row.append(min(ins, delete, replace))
        prev_row = curr_row
    return prev_row[-1]


def similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    dist = levenshtein_distance(a, b)
    return 1.0 - dist / max(len(a), len(b))


def split_values(raw: str) -> List[str]:
    if not raw:
        return []
    parts = re.split(r"[;,|，；、/\s]+", raw.strip())
    return [p for p in parts if p]


def read_excel_rows(excel_path: Path) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "读取 Excel 需要 openpyxl。请先执行: pip install openpyxl"
        ) from exc

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c).strip() if c is not None else "" for c in row])
    if not rows:
        return []

    header = rows[0]
    data_rows = rows[1:]
    result: List[Dict[str, str]] = []
    for row in data_rows:
        item = {}
        for i, name in enumerate(header):
            if not name:
                continue
            item[name] = row[i] if i < len(row) else ""
        result.append(item)
    return result


@dataclass
class Node:
    node_id: str
    intents: List[str]
    slots: List[str] = field(default_factory=list)
    actions: List[str] = field(default_factory=list)
    response: str = ""
    child_nodes: List[str] = field(default_factory=list)


@dataclass
class DialogueState:
    scenario_name: str
    current_node_id: Optional[str]
    active_node_id: Optional[str]
    slots: Dict[str, str] = field(default_factory=dict)
    requested_slot: Optional[str] = None
    is_finished: bool = False


class Scenario:
    def __init__(self, path: Path):
        self.path = path
        data = json.loads(path.read_text(encoding="utf-8"))
        self.nodes: Dict[str, Node] = {}
        for item in data:
            node = Node(
                node_id=item["id"],
                intents=item.get("intent", []),
                slots=item.get("slot", []),
                actions=item.get("action", []),
                response=item.get("response", ""),
                child_nodes=item.get("childnode", []),
            )
            self.nodes[node.node_id] = node
        self.root_node_id = data[0]["id"] if data else None

    def get_node(self, node_id: Optional[str]) -> Optional[Node]:
        if not node_id:
            return None
        return self.nodes.get(node_id)


class SlotOntology:
    def __init__(self):
        self.slot_questions: Dict[str, str] = {}
        self.slot_values: Dict[str, List[str]] = {}

    def load_excel(self, excel_path: Path) -> None:
        rows = read_excel_rows(excel_path)
        if not rows:
            return

        def pick_key(keys: List[str], candidates: List[str]) -> Optional[str]:
            for k in keys:
                normalized = k.strip().lower()
                for c in candidates:
                    if c in normalized:
                        return k
            return None

        keys = list(rows[0].keys())
        slot_key = pick_key(keys, ["槽", "slot"])
        ask_key = pick_key(keys, ["反问", "追问", "ask", "question"])
        value_key = pick_key(keys, ["值", "value", "枚举"])
        if not slot_key:
            raise ValueError("Excel 中未找到槽位列，请确保存在 '槽位' 或 'slot' 相关列名。")

        for row in rows:
            slot = row.get(slot_key, "").strip()
            if not slot:
                continue
            slot_name = slot if slot.startswith("#") else f"#{slot}#"
            if ask_key and row.get(ask_key, "").strip():
                self.slot_questions[slot_name] = row.get(ask_key, "").strip()
            if value_key and row.get(value_key, "").strip():
                self.slot_values[slot_name] = split_values(row.get(value_key, ""))

    def get_question(self, slot_name: str) -> str:
        if slot_name in self.slot_questions:
            return self.slot_questions[slot_name]
        pure_name = slot_name.strip("#")
        return f"请告诉我{pure_name}。"


class NLU:
    def __init__(self, ontology: SlotOntology):
        self.ontology = ontology
        self.default_values = {
            "#支付方式#": ["微信", "支付宝", "银行卡", "信用卡", "现金"],
            "#服装颜色#": ["黑", "白", "红", "蓝", "绿", "灰"],
            "#服装尺寸#": ["S", "M", "L", "XL", "XXL"],
            "#服装类型#": ["衬衫", "外套", "裤子", "卫衣", "毛衣", "T恤"],
        }
        self.ollama_url = "http://localhost:11434/api/generate"
        self.model = "qwen3-14B"

    def _call_ollama(self, prompt: str) -> Optional[str]:
        """调用本地ollama大模型
        
        Args:
            prompt: 提示词
            
        Returns:
            模型的响应，如果调用失败则返回None
        """
        try:
            response = ollama.generate(
                model=self.model,
                prompt=prompt,
                stream=False
            )
            return response.get("response", "").strip()
        except Exception as e:
            # 大模型调用失败，使用传统方法
            return None

    def intent_recognize(self, text: str, candidates: List[Tuple[str, str]]) -> Tuple[Optional[str], float]:
        best_node = None
        best_score = 0.0
        for node_id, intent_text in candidates:
            score = similarity(text, intent_text)
            if intent_text in text:
                score = max(score, 0.95)
            if score > best_score:
                best_score = score
                best_node = node_id
        return best_node, best_score

    def extract_slots(self, text: str, target_slots: List[str]) -> Dict[str, str]:
        # 首先尝试使用大模型提取槽位
        model_result = self._extract_slots_with_model(text, target_slots)
        if model_result:
            return model_result
        
        # 大模型提取失败时，使用传统方法
        result: Dict[str, str] = {}
        for slot in target_slots:
            value = self._extract_one_slot(text, slot)
            if value:
                result[slot] = value
        return result

    def _extract_slots_with_model(self, text: str, target_slots: List[str]) -> Dict[str, str]:
        """使用大模型提取槽位
        
        Args:
            text: 用户输入文本
            target_slots: 目标槽位列表
            
        Returns:
            提取到的槽位值，如果提取失败则返回空字典
        """
        if not target_slots:
            return {}
        
        slot_names = [slot.strip("#") for slot in target_slots]
        prompt = f"从用户输入中提取以下槽位的值：{', '.join(slot_names)}。\n"
        prompt += f"用户输入：{text}\n"
        prompt += "请以JSON格式返回结果，键为槽位名称（不包含#），值为提取到的内容。如果无法提取某个槽位，请不要包含该键。"
        
        response = self._call_ollama(prompt)
        if not response:
            return {}
        
        try:
            # 尝试解析JSON响应
            result = json.loads(response)
            # 将槽位名称转换为带#的格式
            formatted_result = {}
            for slot_name, value in result.items():
                if value:
                    formatted_slot = f"#{slot_name}#"
                    if formatted_slot in target_slots:
                        formatted_result[formatted_slot] = str(value)
            return formatted_result
        except Exception as e:
            print(f"解析大模型响应失败: {e}")
            return {}

    def normalize_free_text_slot_value(self, text: str) -> str:
        cleaned = text.strip()
        cleaned = re.sub(r"^[，。！？、,.!?\s]+|[，。！？、,.!?\s]+$", "", cleaned)
        cleaned = re.sub(r"^(我想要|我要|我想买|买|来个|来件|给我来件|给我|要)\s*", "", cleaned)
        # 处理尺寸中的"号"字，避免重复
        if re.search(r"(\d+|S|M|L|XL|XXL)\s*号", cleaned):
            cleaned = re.sub(r"(\d+|S|M|L|XL|XXL)\s*号", r"\1", cleaned)
        return cleaned.strip()

    def _extract_one_slot(self, text: str, slot: str) -> Optional[str]:
        # 优先使用配置枚举值
        candidates = self.ontology.slot_values.get(slot, []) or self.default_values.get(slot, [])
        for candidate in candidates:
            if candidate and candidate in text:
                return candidate
        if candidates:
            token_candidates = re.findall(r"[\u4e00-\u9fa5A-Za-z0-9]+", text)
            best_value = None
            best_score = 0.0
            for token in token_candidates:
                for candidate in candidates:
                    score = similarity(token.lower(), candidate.lower())
                    if score > best_score:
                        best_score = score
                        best_value = candidate
            if best_score >= 0.6:
                return best_value

        # 通用规则
        if slot == "#分期付款期数#":
            m = re.search(r"(\d+)\s*期", text)
            if m:
                return m.group(1)
            m2 = re.search(r"\b(\d+)\b", text)
            if m2:
                return m2.group(1)
        if slot == "#支付方式#":
            for k in ["微信", "支付宝", "银行卡", "信用卡", "现金"]:
                if k in text:
                    return k
        if slot == "#服装尺寸#":
            m = re.search(r"\b(XXL|XL|L|M|S)\b", text, flags=re.IGNORECASE)
            if m:
                return m.group(1).upper()
            m2 = re.search(r"(\d+)\s*码", text)
            if m2:
                return f"{m2.group(1)}"
        if slot == "#时间#":
            m = re.search(r"(\d{1,2})\s*点", text)
            if m:
                return m.group(1)
        if slot == "#电影名称#":
            m = re.search(r"看(.*?)电影", text)
            if m and m.group(1).strip():
                return m.group(1).strip()
        return None


class DST:
    def update(self, state: DialogueState, recognized_slots: Dict[str, str]) -> DialogueState:
        state.slots.update(recognized_slots)
        return state

    def missing_slots(self, state: DialogueState, node: Node) -> List[str]:
        return [slot for slot in node.slots if slot not in state.slots]


class PM:
    def __init__(self, scenarios: Dict[str, Scenario], nlu: NLU, dst: DST):
        self.scenarios = scenarios
        self.nlu = nlu
        self.dst = dst

    def init_state(self, scenario_name: str) -> DialogueState:
        scenario = self.scenarios[scenario_name]
        return DialogueState(
            scenario_name=scenario_name,
            current_node_id=scenario.root_node_id,
            active_node_id=None,
            slots={},
        )

    def _candidate_nodes(self, state: DialogueState) -> List[str]:
        scenario = self.scenarios[state.scenario_name]
        current = scenario.get_node(state.current_node_id)
        if not current:
            return []
        candidates = [current.node_id]
        candidates.extend(current.child_nodes)
        return candidates

    def step(self, state: DialogueState, user_text: str) -> Tuple[DialogueState, Dict[str, str]]:
        # 检查是否是重听请求
        repeat_keywords = ["重听", "再说一遍", "再说一次", "重复", "再说"]
        for keyword in repeat_keywords:
            if keyword in user_text:
                return state, {"type": "repeat"}
        
        scenario = self.scenarios[state.scenario_name]
        if state.is_finished:
            return state, {"type": "end", "text": "流程已经结束，如需继续请重新开始。"}

        # 检查用户是否在确认当前选择
        confirm_keywords = ["好的", "可以", "没问题", "确认", "是的", "对", "同意"]
        if any(keyword in user_text for keyword in confirm_keywords) and state.active_node_id:
            # 用户确认，尝试跳转到下一个节点
            current_node = scenario.get_node(state.active_node_id)
            if current_node and current_node.child_nodes:
                # 选择第一个子节点作为下一个节点
                next_node_id = current_node.child_nodes[0]
                next_node = scenario.get_node(next_node_id)
                if next_node:
                    state.current_node_id = next_node_id
                    state.active_node_id = next_node_id
                    # 检查下一个节点是否需要槽位
                    missing = self.dst.missing_slots(state, next_node)
                    if missing:
                        state.requested_slot = missing[0]
                        return state, {"type": "ask_slot", "slot": missing[0], "node_id": next_node_id}
                    else:
                        # 下一个节点不需要槽位，直接生成响应
                        if not next_node.child_nodes:
                            state.is_finished = True
                        return state, {"type": "respond", "node_id": next_node_id}

        # 检查用户是否在请求推荐
        recommend_keywords = ["推荐", "有什么", "哪些", "什么", "可选"]
        if any(keyword in user_text for keyword in recommend_keywords) and state.requested_slot:
            # 获取当前请求槽位的可选值
            slot_name = state.requested_slot
            slot_values = self.nlu.ontology.slot_values.get(slot_name, []) or self.nlu.default_values.get(slot_name, [])
            if slot_values:
                # 生成推荐列表
                values_str = "、".join(slot_values)
                return state, {"type": "recommend", "slot": slot_name, "values": slot_values, "text": f"您可以选择：{values_str}。"}

        candidate_node_ids = self._candidate_nodes(state)
        candidate_intents: List[Tuple[str, str]] = []
        for node_id in candidate_node_ids:
            node = scenario.get_node(node_id)
            if not node:
                continue
            for intent_text in node.intents:
                candidate_intents.append((node_id, intent_text))

        target_node_id, score = self.nlu.intent_recognize(user_text, candidate_intents)
        if target_node_id is None or score < 0.45:
            # 用户可能是在补充槽位信息
            target_node_id = state.active_node_id or state.current_node_id

        node = scenario.get_node(target_node_id)
        if not node:
            return state, {"type": "fallback", "text": "抱歉，我没理解你的意思。"}

        state.active_node_id = node.node_id

        extracted = self.nlu.extract_slots(user_text, node.slots)
        # 当系统正在追问某个槽位时，允许用户以自由文本直接回答，避免反复追问
        if (
            state.requested_slot
            and state.requested_slot in node.slots
            and state.requested_slot not in extracted
        ):
            # 检查用户输入是否可能是槽位值
            # 对于数字、颜色、尺寸等类型的槽位，直接接受用户输入
            requested_slot_clean = state.requested_slot.strip("#")
            if any(keyword in requested_slot_clean for keyword in ["期数", "支付", "尺寸", "时间", "颜色"]):
                fallback_value = self.nlu.normalize_free_text_slot_value(user_text)
                if fallback_value:
                    extracted[state.requested_slot] = fallback_value
            else:
                # 对于其他类型的槽位，尝试使用传统方法提取
                fallback_value = self.nlu._extract_one_slot(user_text, state.requested_slot)
                if fallback_value:
                    extracted[state.requested_slot] = fallback_value
        self.dst.update(state, extracted)
        missing = self.dst.missing_slots(state, node)
        if missing:
            state.requested_slot = missing[0]
            return state, {"type": "ask_slot", "slot": missing[0], "node_id": node.node_id}

        state.requested_slot = None
        state.current_node_id = node.node_id
        if not node.child_nodes:
            state.is_finished = True
        return state, {"type": "respond", "node_id": node.node_id}


class NLG:
    def __init__(self, ontology: SlotOntology):
        self.ontology = ontology
        self.last_response = ""  # 存储上一次的响应

    def render(self, state: DialogueState, scenario: Scenario, policy_output: Dict[str, str]) -> str:
        # 检查是否是重听请求
        if policy_output.get("type") == "repeat":
            return self.last_response or "抱歉，我还没有说过什么。"
        
        output_type = policy_output.get("type")
        if output_type == "end":
            response = policy_output.get("text", "流程结束。")
        elif output_type == "fallback":
            response = policy_output.get("text", "抱歉，我不太明白。")
        elif output_type == "ask_slot":
            slot = policy_output["slot"]
            response = self.ontology.get_question(slot)
        elif output_type == "recommend":
            response = policy_output.get("text", "")
        else:
            node = scenario.get_node(policy_output.get("node_id"))
            if not node:
                response = "抱歉，响应生成失败。"
            else:
                text = node.response
                for slot_name, slot_value in state.slots.items():
                    text = text.replace(slot_name, str(slot_value))
                if node.actions:
                    actions_text = "；".join(node.actions)
                    response = f"{text}\n[执行动作] {actions_text}"
                else:
                    response = text
        
        # 保存当前响应
        self.last_response = response
        return response


class DialogueSystem:
    def __init__(self, scenario_dir: Path, excel_path: Optional[Path] = None):
        self.scenario_dir = scenario_dir
        self.scenarios = self._load_scenarios(scenario_dir)
        if not self.scenarios:
            raise RuntimeError(f"未在 {scenario_dir} 中找到场景 JSON。")
        self.ontology = SlotOntology()
        if excel_path and excel_path.exists():
            self.ontology.load_excel(excel_path)
        self.nlu = NLU(self.ontology)
        self.dst = DST()
        self.pm = PM(self.scenarios, self.nlu, self.dst)
        self.nlg = NLG(self.ontology)
        self.state: Optional[DialogueState] = None

    def _load_scenarios(self, scenario_dir: Path) -> Dict[str, Scenario]:
        scenarios = {}
        for p in sorted(scenario_dir.glob("*.json")):
            scenario_name = p.stem.replace("scenario-", "")
            scenarios[scenario_name] = Scenario(p)
        return scenarios

    def list_scenarios(self) -> List[str]:
        return list(self.scenarios.keys())

    def start(self, scenario_name: str) -> None:
        if scenario_name not in self.scenarios:
            raise ValueError(f"未知场景: {scenario_name}")
        self.state = self.pm.init_state(scenario_name)

    def chat(self, user_text: str) -> str:
        if not self.state:
            return "请先选择并启动一个场景。"
        state, pm_output = self.pm.step(self.state, user_text)
        self.state = state
        scenario = self.scenarios[self.state.scenario_name]
        return self.nlg.render(self.state, scenario, pm_output)


def main() -> None:
    base_dir = Path(__file__).parent
    excel_candidates = [
        base_dir / "slot_fitting_templet.xlsx",
        base_dir / "slot_ontology.xlsx",
        base_dir / "slot_ontology.xls",
    ]
    excel_path = next((p for p in excel_candidates if p.exists()), None)
    ds = DialogueSystem(base_dir, excel_path=excel_path)

    scenarios = ds.list_scenarios()
    print("可用场景：")
    for i, name in enumerate(scenarios, start=1):
        print(f"{i}. {name}")
    choice = input("请输入场景编号: ").strip()
    if not choice.isdigit() or not (1 <= int(choice) <= len(scenarios)):
        print("无效编号，程序结束。")
        return
    scenario_name = scenarios[int(choice) - 1]
    ds.start(scenario_name)

    print(f"已进入场景：{scenario_name}")
    print("输入 quit 退出。")
    print("输入 重听/再说一遍 等关键词可重复上一条回答。")
    while True:
        user_text = input("你: ").strip()
        if user_text.lower() in {"quit", "exit"}:
            print("系统: 再见。")
            break
        bot = ds.chat(user_text)
        print(f"系统: {bot}")


if __name__ == "__main__":
    main()
